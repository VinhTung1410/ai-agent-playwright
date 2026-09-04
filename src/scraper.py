import os
import re
import time
import logging
import pandas as pd
from urllib.parse import urlparse
from playwright.sync_api import sync_playwright
import config

# openpyxl styles and tools
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# Standard skill dictionary in French
SKILL_MAP = {
    # Technical tools & software -> Outil Technique / Logiciel
    "SQL": (["sql"], "Outil Technique / Logiciel"),
    "Power BI": (["power bi", "powerbi"], "Outil Technique / Logiciel"),
    "Excel": (["excel", "xlsx"], "Outil Technique / Logiciel"),
    "Jira": (["jira"], "Outil Technique / Logiciel"),
    "Python": (["python"], "Outil Technique / Logiciel"),
    "Tableau": (["tableau"], "Outil Technique / Logiciel"),
    "UML": (["uml"], "Outil Technique / Logiciel"),
    "BPMN": (["bpmn"], "Outil Technique / Logiciel"),
    "SAP": (["sap"], "Outil Technique / Logiciel"),
    "Salesforce": (["salesforce"], "Outil Technique / Logiciel"),
    "Confluence": (["confluence"], "Outil Technique / Logiciel"),
    "MS Visio": (["visio"], "Outil Technique / Logiciel"),
    "PowerPoint": (["powerpoint", "power point"], "Outil Technique / Logiciel"),
    "SharePoint": (["sharepoint"], "Outil Technique / Logiciel"),
    "Snowflake": (["snowflake"], "Outil Technique / Logiciel"),
    "Alteryx": (["alteryx"], "Outil Technique / Logiciel"),
    "R": (["\\br\\b"], "Outil Technique / Logiciel"),
    
    # Methodologies & frameworks -> Méthodologie / Framework
    "Agile": (["agile"], "Méthodologie / Framework"),
    "Scrum": (["scrum"], "Méthodologie / Framework"),
    "Kanban": (["kanban"], "Méthodologie / Framework"),
    "Waterfall (Cycle en V)": (["waterfall", "cycle en v", "cycle en-v"], "Méthodologie / Framework"),
    "SAFe": (["safe"], "Méthodologie / Framework"),
    "DevOps": (["devops"], "Méthodologie / Framework"),
    
    # Soft skills & business competencies -> Compétence Professionnelle / Soft Skill
    "Gestion des parties prenantes": (["stakeholder", "parties prenantes", "relation client"], "Compétence Professionnelle / Soft Skill"),
    "Recueil des besoins": (["requirement", "besoin", "cahier des charges", "spécification", "user story", "user stories"], "Compétence Professionnelle / Soft Skill"),
    "Gestion du changement": (["change management", "conduite du changement"], "Compétence Professionnelle / Soft Skill"),
    "Analyse de données": (["data analysis", "analyse de données", "analyse des données"], "Compétence Professionnelle / Soft Skill"),
    "Gestion de projet": (["project management", "gestion de projet", "gestion de projets"], "Compétence Professionnelle / Soft Skill"),
    "Esprit d'analyse": (["analytical", "esprit d'analyse", "esprit d’analyse", "capacités d'analyse"], "Compétence Professionnelle / Soft Skill"),
    "Compétences en communication": (["communication", "aisance relationnelle"], "Compétence Professionnelle / Soft Skill"),
    "Résolution de problèmes": (["problem solving", "résolution de problèmes"], "Compétence Professionnelle / Soft Skill"),
    
    # Language requirements -> Langue
    "Français": (["french", "français", "francais"], "Langue"),
    "Anglais": (["english", "anglais"], "Langue"),
}

def safe_screenshot(page, path, timeout=3500):
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Bypass Playwright font lock on LinkedIn
        try:
            page.evaluate("""
                try {
                    Object.defineProperty(document, 'fonts', {
                        value: { ready: Promise.resolve(), status: 'loaded', check: () => true }
                    });
                } catch(e) {}
            """)
        except Exception:
            pass

        # Hide any obstructive headers, cookie banners, or auth modals
        try:
            page.evaluate("""
                const hideStyles = document.createElement('style');
                hideStyles.innerHTML = `
                    .modal__overlay, .top-level-modal-container, .contextual-sign-in-modal,
                    .authwall, #advocate-modal, .artdeco-global-alert-container,
                    #artdeco-global-alert-container, nav.nav, footer, .cookie-banner { 
                        display: none !important; 
                    }
                `;
                document.head.appendChild(hideStyles);
            """)
        except Exception:
            pass

        # Target specifically the company and job header card (top-card)
        top_card = page.locator(".top-card-layout, section.top-card-layout, .top-card-layout__card, .topcard, div[class*='top-card']").first
        if top_card.count() > 0 and top_card.is_visible():
            top_card.screenshot(path=path, timeout=timeout, animations="disabled")
        else:
            # Fallback to job container if top-card is not found
            job_container = page.locator(".core-section-container, .decorated-job-posting__details, main.main-content, .job-view-layout").first
            if job_container.count() > 0 and job_container.is_visible():
                job_container.screenshot(path=path, timeout=timeout, animations="disabled")
            else:
                page.screenshot(path=path, full_page=False, timeout=timeout, animations="disabled")
    except Exception as e:
        logger.warning(f"Container/Full screenshot failed ({e}), attempting standard screenshot...")
        try:
            page.screenshot(path=path, full_page=False, timeout=1500, animations="disabled")
        except Exception as e2:
            logger.warning(f"Screenshot capture failed: {e2}")

def clean_filename(name):
    name = re.sub(r'[\\/*?:"<>| ]', '_', name)
    name = re.sub(r'_+', '_', name)
    return name.strip('_')

def extract_skills_from_text(text):
    if not text:
        return []
    extracted = []
    text_lower = text.lower()
    for skill_name, (patterns, category) in SKILL_MAP.items():
        found = False
        for pattern in patterns:
            if re.search(pattern, text_lower):
                found = True
                break
        if found:
            extracted.append(skill_name)
    return extracted

def scrape_linkedin(keywords="alternance business analyst", location="France", max_jobs=10, progress_callback=None):
    os.makedirs("output/screenshots", exist_ok=True)

    if progress_callback:
        progress_callback(0, max_jobs, "Démarrage du navigateur Chromium...")

    with sync_playwright() as p:
        # Launch standard Chromium (headless on cloud/Linux, headful on local)
        import sys
        is_cloud = (
            os.environ.get("STREAMLIT_SERVER_PORT") is not None
            or os.environ.get("PORT") is not None
            or os.environ.get("IS_STREAMLIT_CLOUD") is not None
            or sys.platform != "win32"
        )
        
        headless_mode = True if is_cloud else False
        logger.info(f"Launching Chromium (headless={headless_mode})...")
        browser = p.chromium.launch(headless=headless_mode)
                
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        
        # Block heavy telemetry, beacons, media and third-party trackers to cut page load time
        def block_unnecessary_resources(route):
            req = route.request
            url = req.url
            rtype = req.resource_type
            if rtype in ["media", "beacon", "websocket"] or any(tracker in url for tracker in [
                "google-analytics", "analytics", "doubleclick", "scorecardresearch",
                "linkedin.com/li/track", "platform.linkedin.com/litms", "telemetry"
            ]):
                route.abort()
            else:
                route.continue_()
                
        context.route("**/*", block_unnecessary_resources)

        page = context.new_page()
        page.set_default_timeout(15000)

        # URL encode keyword and location
        from urllib.parse import quote
        encoded_keywords = quote(keywords)
        encoded_location = quote(location)
        search_url = f"https://www.linkedin.com/jobs/search?keywords={encoded_keywords}&location={encoded_location}&f_TPR=r86400"
        logger.info(f"Navigating to: {search_url}")
        
        if progress_callback:
            progress_callback(0, max_jobs, f"Recherche sur LinkedIn pour '{keywords}' ({location})...")
            
        page.goto(search_url, wait_until="domcontentloaded", timeout=15000)

        # Inject CSS to hide obstructive login modals and overlays
        try:
            page.evaluate("""
                const styles = document.createElement('style');
                styles.innerHTML = `
                    .modal__overlay, 
                    .top-level-modal-container, 
                    .contextual-sign-in-modal,
                    .authwall,
                    #advocate-modal { 
                        display: none !important; 
                        pointer-events: none !important; 
                    }
                    body {
                        overflow: auto !important;
                    }
                `;
                document.head.appendChild(styles);
            """)
            logger.info("Injected CSS styles to hide modal overlays.")
        except Exception as e:
            logger.warning(f"Could not inject CSS overlay style: {e}")

        # Accept cookies if visible
        try:
            accept_btn = page.locator("button:has-text('Accept'), button:has-text('Accepter'), button.artdeco-global-alert__action").first
            if accept_btn.count() > 0 and accept_btn.is_visible():
                accept_btn.click(timeout=1000)
                logger.info("Dismissed cookie consent banner.")
        except Exception as e:
            logger.warning(f"Could not dismiss cookie banner: {e}")

        # Wait for the job cards to appear
        try:
            page.wait_for_selector(".base-card, .base-search-card, a.base-card__full-link", timeout=6000)
            logger.info("Job cards loaded.")
        except Exception as e:
            logger.warning(f"Timeout waiting for job cards: {e}")
            safe_screenshot(page, "output/screenshots/search_page_error.png")

        # Initial check for job card links
        def get_current_card_urls():
            found_urls = []
            cards = page.locator("a.base-card__full-link, a.job-search-card__image-link").all()
            for card in cards:
                href = card.get_attribute("href")
                if href:
                    clean_url = href.split("?")[0]
                    if clean_url not in found_urls:
                        found_urls.append(clean_url)
            return found_urls

        raw_urls = get_current_card_urls()

        # Only scroll if we need more job listings
        if len(raw_urls) < max_jobs:
            scroll_times = min(4, ((max_jobs - len(raw_urls)) // 4) + 1)
            logger.info(f"Scrolling {scroll_times} times to reach {max_jobs} job listings...")
            
            if progress_callback:
                progress_callback(0, max_jobs, f"Chargement de la liste des offres ({scroll_times} défilements)...")
                
            for _ in range(scroll_times):
                page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                page.wait_for_timeout(800)
                see_more = page.locator("button.infinite-scroller__show-more-button")
                if see_more.is_visible():
                    try:
                        see_more.click(timeout=800)
                        page.wait_for_timeout(800)
                    except Exception as e:
                        logger.warning(f"See more button click failed: {e}")
                raw_urls = get_current_card_urls()
                if len(raw_urls) >= max_jobs:
                    break

        target_count = min(max_jobs, len(raw_urls))
        if target_count < max_jobs:
            logger.warning(f"Found only {len(raw_urls)} jobs (requested {max_jobs}). Processing all available.")
            target_count = len(raw_urls)
            
        logger.info(f"Targeting {target_count} jobs out of {len(raw_urls)} found.")
        
        jobs_data = []
        
        for idx, job_url in enumerate(raw_urls[:target_count], start=1):
            logger.info(f"Processing job {idx}/{target_count}: {job_url}")
            if progress_callback:
                progress_callback(idx - 1, target_count, f"Extraction [{idx}/{target_count}] : Chargement de la page...")

            try:
                page.goto(job_url, wait_until="domcontentloaded", timeout=10000)
                
                # Fast wait for core elements to appear (DOM is already loaded)
                try:
                    page.wait_for_selector(".description__text, .show-more-less-html__markup, h1", timeout=3000)
                except Exception:
                    pass

                # Dismiss login modal if visible
                dismiss_btn = page.locator("button.modal__dismiss").first
                if dismiss_btn.count() > 0 and dismiss_btn.is_visible():
                    try:
                        dismiss_btn.click(timeout=600)
                    except Exception:
                        pass

                # Extract Job Title
                title_sel = page.locator("h1.top-card-layout__title, h1.topcard__title, h1")
                job_title = "Unknown Title"
                if title_sel.count() > 0:
                    job_title = title_sel.first.text_content().strip()

                # Company & Link
                comp_sel = page.locator("a.topcard__org-name-link, span.topcard__flavor, a[href*='/company/']")
                company = "Unknown Company"
                company_url = ""
                if comp_sel.count() > 0:
                    company = comp_sel.first.text_content().strip()
                    href = comp_sel.first.get_attribute("href") or ""
                    if href:
                        company_url = href.split("?")[0]
                        if company_url.startswith("/"):
                            company_url = f"https://www.linkedin.com{company_url}"

                if progress_callback:
                    progress_callback(idx - 1, target_count, f"Extraction [{idx}/{target_count}] : {company} - {job_title[:30]}...")

                # Location
                loc_sel = page.locator("span.topcard__flavor--metadata, span.sub-nav-item__sub-text")
                location = "France"
                if loc_sel.count() > 0:
                    location = loc_sel.first.text_content().strip()

                # Criteria (Employment type, Seniority, Industries, Job function)
                job_criteria = {}
                try:
                    for item in page.locator("li.description__job-criteria-item").all():
                        h = item.locator("h3").first.text_content().strip() if item.locator("h3").count() > 0 else ""
                        v = item.locator("span").first.text_content().strip() if item.locator("span").count() > 0 else ""
                        if h and v:
                            job_criteria[h] = v
                except Exception:
                    pass

                emp_type = job_criteria.get("Type d’emploi") or job_criteria.get("Employment type") or "Alternance / CDI"
                seniority = job_criteria.get("Niveau hiérarchique") or job_criteria.get("Seniority level") or "Non spécifié"
                industries = job_criteria.get("Secteurs") or job_criteria.get("Industries") or "Non spécifié"
                job_fn = job_criteria.get("Fonction") or job_criteria.get("Job function") or "Non spécifié"

                # Description (About the job)
                desc_sel = page.locator(".description__text, .show-more-less-html__markup")
                description = ""
                if desc_sel.count() > 0:
                    show_more = page.locator("button.show-more-less-html__button").first
                    if show_more.count() > 0 and show_more.is_visible():
                        try:
                            show_more.click(timeout=600)
                            page.wait_for_timeout(200)
                        except Exception:
                            pass
                    description = desc_sel.first.text_content().strip()

                # Skills
                skills_required = extract_skills_from_text(description)
                
                jobs_data.append({
                    "ID": idx,
                    "Job Title": job_title,
                    "Company": company,
                    "Location": location,
                    "Employment Type": emp_type,
                    "Seniority Level": seniority,
                    "Industries": industries,
                    "Job Function": job_fn,
                    "Company URL": company_url,
                    "Job URL": job_url,
                    "Description": description,
                    "Key Skills Required": ", ".join(skills_required),
                    "Skills List": skills_required
                })
                
                if progress_callback:
                    progress_callback(idx, target_count, f"Terminé [{idx}/{target_count}] : {company} ({len(skills_required)} compétences trouvées)")
                
                time.sleep(0.05)

            except Exception as e:
                logger.error(f"Error processing job {idx}: {e}")
                continue

        if progress_callback:
            progress_callback(target_count, target_count, "Analyse terminée. Fermeture du navigateur...")

        browser.close()
        
    return jobs_data

def process_and_export(jobs_data):
    if not jobs_data:
        logger.error("No jobs data scraped.")
        print("No jobs data scraped.")
        return
        
    df_jobs = pd.DataFrame(jobs_data)
    
    total_jobs = len(df_jobs)
    skill_counts = {}
    
    for row in jobs_data:
        for skill in row["Skills List"]:
            if skill not in skill_counts:
                skill_counts[skill] = 0
            skill_counts[skill] += 1
            
    skill_ranking = []
    for skill, count in skill_counts.items():
        category = SKILL_MAP[skill][1]
        rate = (count / total_jobs) * 100.0
        skill_ranking.append({
            "Skill Name": skill,
            "Category": category,
            "Job Count": count,
            "Occurrence Rate (%)": f"{rate:.1f}%"
        })
        
    df_ranking = pd.DataFrame(skill_ranking)
    if not df_ranking.empty:
        df_ranking = df_ranking.sort_values(by="Job Count", ascending=False).reset_index(drop=True)
        df_ranking.insert(0, "Rank", df_ranking.index + 1)
        df_ranking["Rank"] = df_ranking["Rank"].apply(lambda r: f"Top {r}")
    else:
        df_ranking = pd.DataFrame(columns=["Rank", "Skill Name", "Category", "Job Count", "Occurrence Rate (%)"])
        
    # Prepare df_jobs for Sheet 1
    sheet_cols = [
        "ID", "Job Title", "Company", "Location", 
        "Employment Type", "Seniority Level", "Industries",
        "Job URL", "Key Skills Required"
    ]
    available_cols = [c for c in sheet_cols if c in df_jobs.columns]
    df_jobs_sheet = df_jobs[available_cols].copy()
    
    # Add application tracker columns to help searchers manage their process
    df_jobs_sheet["Statut Candidature"] = "À postuler"
    df_jobs_sheet["Date Candidature"] = ""
    df_jobs_sheet["Contact Recruteur"] = ""
    df_jobs_sheet["Notes & Remarques"] = ""
    if "Description" in df_jobs.columns:
        df_jobs_sheet["Description"] = df_jobs["Description"]
    
    # Save to Excel
    excel_path = "output/LinkedIn_BA_France_Report.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    
    num_cols = len(df_jobs_sheet.columns)
    
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_jobs_sheet.to_excel(writer, sheet_name="Job Listings", index=False)
        df_ranking.to_excel(writer, sheet_name="Top Skills Ranking", index=False)
        
        workbook = writer.book
        
        # --- Format Sheet 1: Job Listings ---
        ws_jobs = writer.sheets["Job Listings"]
        ws_jobs.views.sheetView[0].showGridLines = True
        
        # Styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Elegant Dark Blue
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Add Dropdown Data Validation for Statut Candidature (Column J)
        dv = DataValidation(type="list", formula1='"À postuler,Postulé,Entretien,Refusé,Offre reçue"', allow_blank=True)
        dv.error ='Saisie non valide'
        dv.errorTitle = 'Erreur'
        dv.prompt = 'Veuillez choisir un statut dans la liste'
        dv.promptTitle = 'Statut Candidature'
        ws_jobs.add_data_validation(dv)
        dv.add(f"J2:J{total_jobs + 1}")
        
        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Format Headers
        for col_idx in range(1, num_cols + 1):
            cell = ws_jobs.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Format Rows
        for r_idx in range(2, total_jobs + 2):
            for c_idx in range(1, num_cols + 1):
                cell = ws_jobs.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                
                if c_idx in [1, 4, 5, 6, 10, 11]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                
                # Clickable Job URL hyperlink (Column 8)
                if c_idx == 8 and cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")
                
                # Clickable Screenshot File hyperlink
                if c_idx == 7 and cell.value:
                    abs_path = os.path.abspath(cell.value)
                    cell.hyperlink = abs_path
                    cell.font = Font(color="0563C1", underline="single")
                    
        # Auto-adjust Column Widths for Job Listings
        for col in ws_jobs.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > 40:
                    val_str = val_str[:40]
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_jobs.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # --- Format Sheet 2: Top Skills Ranking ---
        ws_ranking = writer.sheets["Top Skills Ranking"]
        ws_ranking.views.sheetView[0].showGridLines = True
        
        # Format Headers
        for col_idx in range(1, 6):
            cell = ws_ranking.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Format Rows
        for r_idx in range(2, len(df_ranking) + 2):
            for c_idx in range(1, 6):
                cell = ws_ranking.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if c_idx in [1, 4, 5]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
        # Auto-adjust Column Widths for Top Skills Ranking
        for col in ws_ranking.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws_ranking.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    logger.info(f"Excel report saved successfully to {excel_path}")
    print(f"Excel report saved successfully to {excel_path}")
